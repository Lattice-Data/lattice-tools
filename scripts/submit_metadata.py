import argparse
import ast
import os
import json
import lattice
import sys
import logging
import mimetypes
import requests
import magic  # install me with 'pip install python-magic'
from PIL import Image  # install me with 'pip install Pillow'
from openpyxl import load_workbook
from urllib.parse import urljoin, quote
from base64 import b64encode


EPILOG = '''
Thorough instructions have been documented in lattice-tools/docs/submit_metadata.md

For more details:

		%(prog)s --help
'''


ORDER = [
    'user',
    'award',
    'lab',
    'organism',
    'gene',
    'publication',
    'document',
    'ontology_term',
    'library_protocol',
    'target',
    'antibody_lot',
    'treatment',
    'human_postnatal_donor',
    'human_prenatal_donor',
    'mouse_postnatal_donor',
    'mouse_prenatal_donor',
    'tissue',
    'cell_culture',
    'organoid',
    'suspension',
    'tissue_section',
    'dataset',
    'library',
    'sequencing_run',
    'raw_sequence_file',
    'sequence_alignment_file',
    'matrix_file',
    'raw_matrix_file',
    'processed_matrix_file',
    'rna_metrics',
    'antibody_capture_metrics',
    'rna_aggregate_metrics',
    'atac_metrics',
    'atac_aggregate_metrics',
    'cell_annotation',
    'image',
    'page',
    'access_key'
]


def getArgs():
	parser = argparse.ArgumentParser(
		description=__doc__, epilog=EPILOG,
		formatter_class=argparse.RawDescriptionHelpFormatter,
	)

	parser.add_argument('infile',
						help='the datafile containing object data to import')
	parser.add_argument('--justtype', '-jt',
						help='the type of the objects to import if only one type is desired')
	parser.add_argument('--starttype', '-st',
						help='the type of the objects to start importing and continue in sequence')
	parser.add_argument('--mode', '-m',
						help='The machine to run on.')
	parser.add_argument('--debug', '-d',
						default=False,
						action='store_true',
						help='Print debug messages.  Default is False')
	parser.add_argument('--update',
						default=False,
						action='store_true',
						help='Let the script proceed with the changes.  Default is False'),
	parser.add_argument('--patchall', '-p',
						default=False,
						action='store_true',
						help='PATCH existing objects.  Default is False \
						and will only PATCH with user override')
	parser.add_argument('--remove',
						default=False,
						action='store_true',
						help='Will remove all values in the provided properties.  Default is False'),
	args = parser.parse_args()
	return args


def reader(book, sheetname):
	sheet = book[sheetname]
	row_count = 1 # start with 1 for the headers
	rows = []
	for row in sheet.iter_rows():
		row = ['' if v.value is None else cell_value(v) for v in row]
		if len(row) != row.count(''):
			if row[0] != '' and str(row[0])[0] == '#': # avoid rows starting with comments
				row_count += 1 # but add one row for each, assuming they are above submission rows
			else:
				rows.append(list(row))
	cell_A1 = rows[0][0]
	if cell_A1 is None or cell_A1.startswith('schema_version='): # may need to delete values in the first column
		for row in rows:
			del row[0]
	# now remove the values in empty columns
	bad_elements = []
	for i in range(0,len(rows[0])):
		temp = []
		for row in rows:
			temp.append(row[i])
		if len(temp) == temp.count(''):
			bad_elements.append(i)
	for ele in sorted(bad_elements, reverse = True):
		for row in rows:
			del row[ele]
	return(row_count, rows)


def cell_value(cell):
	ctype = cell.data_type
	cvalue = cell.value

	if ctype in ['s','n']: # TYPE_STRING, TYPE_NUMERIC
		return cvalue

	elif ctype == 'b': # TYPE_BOOL
		return str(cvalue).upper()

	elif ctype == 'd': # datetime
		return cvalue.date()

	elif ctype == 'f': # TYPE_FORMULA
		raise ValueError(repr(cell), 'formula in cell')

	elif ctype == 'e': # TYPE_ERROR
		raise ValueError(repr(cell), 'cell error')

	else:
		raise ValueError(repr(cell), '-'.join(['unknown cell type',ctype,cvalue]))


def properties_validator(keys, schema, schema_properties, ont_props):
	flag = False
	dup_keys = []
	for key in keys:
		if key is not None:
			if keys.count(key) >1 and key not in dup_keys: # check for duplicated headers
				print('Property {} found {} times in {} sheet headers'.format(key, keys.count(key), schema))
				dup_keys.append(key)
				flag = True
			props = key.split('.')
			propA = props[0].split('-')[0]
			if propA not in schema_properties.keys():
				print('Property {} not found in schema {}'.format(propA, schema))
				flag = True
			elif len(props) > 1:
				propB = props[1].split('-')[0]
				if schema_properties[propA]['type'] == 'array':
					if propB not in schema_properties[propA]['items']['properties'].keys():
						print('Property {} not found in schema {}.{}'.format(propB, schema, propA))
						flag = True
					elif len(props) > 2:
						propC = props[2].split('-')[0]
						if schema_properties[propA]['items']['properties'][propB]['type'] == 'array':
							if propC not in schema_properties[propA]['items']['properties'][propB]['items']['properties'].keys():
								print('Property {} not found in schema {}.{}.{}'.format(propC, schema, propA, propB))
								flag = True
						elif schema_properties[propA]['items']['properties'][propB]['type'] == 'object':
							if propB not in schema_properties[propA]['items']['properties'][propB]['properties'].keys():
								print('Property {} not found in schema {}.{}.{}'.format(propC, schema, propA, propB))
								flag = True
						elif schema_properties[propA]['items']['properties'][propB].get('linkTo') == "OntologyTerm":
							if propC not in ont_props:
								print('Property {} not found in OntologyTerm schema via schema {}.{}.{}'.format(propC, schema, propA, propB))
								flag = True
						else:
							print('Not expecting subproperties for property {}.{} in schema {} ({} given) '.format(propA, propB, schema, propC))
							flag = True
				elif schema_properties[propA]['type'] == 'object':
					if propB not in schema_properties[propA]['properties'].keys() and schema_properties[propA].get('additionalProperties') != True:
						print('Property {} not found in schema {}.{}'.format(propB, schema, propA))
						flag = True
					elif len(props) > 2:
						propC = props[2].split('-')[0]
						if schema_properties[propA]['properties'][propB]['type'] == 'array':
							if propC not in schema_properties[propA]['properties'][propB]['items']['properties'].keys():
								print('Property {} not found in schema {}.{}.{}'.format(propC, schema, propA, propB))
								flag = True
						elif schema_properties[propA]['properties'][propB]['type'] == 'object':
							if propC not in schema_properties[propA]['properties'][propB]['properties'].keys():
								print('Property {} not found in schema {}.{}.{}'.format(propC, schema, propA, propB))
								flag = True
						else:
							print('Not expecting subproperties for property {} in schema {}.{} ({} given) '.format(probB, schema, propA, propC))
							flag = True
				elif schema_properties[propA].get('linkTo') == "OntologyTerm":
					if propB not in ont_props:
						print('Property {} not found in OntologyTerm schema via schema {}.{}'.format(propB, schema, propA))
						flag = True
				else:
					print('Not expecting subproperties for property {} in schema {} ({} given) '.format(propA, schema, propB))
					flag = True
	return flag


def type_formatter(old_value, schema_properties, key1, key2=None, key3=None):
	# determine the type specified in the schema
	linkTo_flag = False
	desired_type = ''
	if not key2:
		desired_type = schema_properties[key1]['type']
		if schema_properties[key1].get('linkTo'):
			linkTo_flag = True
		if schema_properties[key1].get('items'):
			array_of_type = schema_properties[key1]['items'].get('type')
			if schema_properties[key1]['items'].get('linkTo'):
				linkTo_flag = True
	elif not key3:
		if schema_properties[key1].get('items'):
			if schema_properties[key1]['items']['properties'].get(key2):
				desired_type = schema_properties[key1]['items']['properties'][key2]['type']
				if schema_properties[key1]['items']['properties'][key2].get('linkTo'):
					linkTo_flag = True
				if schema_properties[key1]['items']['properties'][key2].get('items'):
					array_of_type = schema_properties[key1]['items']['properties'][key2]['items'].get('type')
		else:
			if schema_properties[key1]['properties'].get(key2):
				desired_type = schema_properties[key1]['properties'][key2]['type']
				if schema_properties[key1]['properties'][key2].get('linkTo'):
					linkTo_flag = True
	else:
		if schema_properties[key1].get('items'):
			if schema_properties[key1]['items']['properties'].get(key2):
				if schema_properties[key1]['items']['properties'][key2].get('items'):
					if schema_properties[key1]['items']['properties'][key2]['items']['properties'].get(key3):
						desired_type = schema_properties[key1]['items']['properties'][key2]['items']['properties'][key3]['type']
						if schema_properties[key1]['items']['properties'][key2]['items']['properties'][key3].get('linkTo'):
							linkTo_flag = True
				else:
					if schema_properties[key1]['items']['properties'][key2]['properties'].get(key3):
						desired_type = schema_properties[key1]['items']['properties'][key2]['properties'][key3]['type']
						if schema_properties[key1]['items']['properties'][key2]['properties'][key3].get('linkTo'):
							linkTo_flag = True
		else:
			if schema_properties[key1]['properties'].get(key2):
				if schema_properties[key1]['properties'][key2].get('items'):
					desired_type = schema_properties[key1]['properties'][key2]['items']['properties'][key3]['type']
					if schema_properties[key1]['properties'][key2]['items']['properties'][key3].get('linkTo'):
						linkTo_flag = True
				else:
					desired_type = schema_properties[key1]['properties'][key2]['properties'][key3]['type']
					if schema_properties[key1]['properties'][key2]['properties'][key3].get('linkTo'):
						linkTo_flag = True

	# adjust the value to the specified type
	if desired_type == 'array' and linkTo_flag == True:
		return [quote(x.strip()) for x in old_value.split(',')]
	if desired_type == 'array':
		if array_of_type == 'object':
			old_value = old_value.replace('},{', '}${')
			return [ast.literal_eval(x) for x in old_value.split('$')]
		else:
			return [x.strip() for x in old_value.split(',')]
	elif desired_type == 'boolean':
		if str(old_value).lower() in ['true', '1']:
			return True
		elif str(old_value).lower() in ['false', '0']:
			return False
		else:
			print('Boolean was expected for {prop} but value is {value}'.format(prop=key1, value=old_value))
	elif desired_type == 'number':
		return float(old_value)
	elif desired_type == 'integer':
		return int(old_value)
	elif desired_type == 'string' and linkTo_flag == True:
		return quote(str(old_value).strip())
	elif desired_type == 'string':
		return str(old_value).strip()
	else:
		return old_value


def dict_patcher(old_dict, schema_properties, ont_schema_properties):
	new_dict = {}
	array_o_objs_dict = {}
	onts = {}
	for key in sorted(old_dict.keys()):
		if old_dict[key] != '':  # this removes empty cells
			path = key.split('.')
			if len(path) == 1:
				if schema_properties[key].get('linkTo') == 'OntologyTerm' or \
					(schema_properties[key].get('items') and schema_properties[key]['items'].get('linkTo') == 'OntologyTerm'):
					old_dict[key] = old_dict[key].replace(':','_')
				new_dict[key] = type_formatter(old_dict[key], schema_properties, key)
			elif len(path) == 2: # embedded object, need to build the mini dictionary to put this in
				if '-' in path[0]: # this has a number next to it and we expect an array of objects
					group = path[0].split('-')[1]
				else:
					group = '1'
				propA = path[0].split('-')[0]
				propB = path[1]
				if schema_properties[propA].get('linkTo') == 'OntologyTerm':
					if onts.get(propA):
						onts[propA][propB] = type_formatter(old_dict[key], ont_schema_properties, propB)
					else:
						onts[propA] = {}
						onts[propA][propB] = type_formatter(old_dict[key], ont_schema_properties, propB)
					if propB == 'term_id':
						new_dict[propA] = old_dict[key].replace(':','_') # replace the term_name/term_id with the linkTo
				elif schema_properties[propA]['type'] == 'array': # this is a single object that needs to end as a 1 item array of objects
					if array_o_objs_dict.get(propA): # I have already added the embedded object to the new dictionary
						if array_o_objs_dict[propA].get(group):
							# this should be that we have started putting in the new object
							array_o_objs_dict[propA][group].update({propB: type_formatter(old_dict[key], schema_properties, propA, propB)})
						else:
							# this means we have not added any part of new item to the list
							array_o_objs_dict[propA][group] = {propB: type_formatter(old_dict[key], schema_properties, propA, propB)}
					else: # I need to make new item in dictionary
						array_o_objs_dict[propA] = {}
						array_o_objs_dict[propA][group] = {propB: type_formatter(old_dict[key], schema_properties, propA, propB)}
				else: # there is no number next to it, this is our only object to handle
					if new_dict.get(propA): # I have already added the embedded object to the new dictionary
						new_dict[propA].update({propB: type_formatter(old_dict[key], schema_properties, propA, propB)})
					else: # I need to make new item in dictionary
						temp_dict = {propB: type_formatter(old_dict[key], schema_properties, propA, propB)}
						new_dict[propA] = temp_dict
			elif len(path) == 3: # we have an object embedded within an embedded object
				if '-' in path[0]:
					group = path[0].split('-')[1]
				else:
					group = '1'
				if '-' in path[1]:
					subgroup = path[1].split('-')[1]
				else:
					subgroup = '1'
				propA = path[0].split('-')[0]
				propB = path[1].split('-')[0]
				propC = path[2]
				if schema_properties[propA]['type'] == 'array': # this is a single object that needs to end as a 1 item array of objects
					if schema_properties[propA]['items']['properties'][propB]['type'] == 'array': # we have an array of objects within an array of objects
						if array_o_objs_dict.get(propA):
							if array_o_objs_dict[propA].get(group):
								if array_o_objs_dict[propA][group].get(propB):
									if array_o_objs_dict[propA][group][propB].get(subgroup):
										array_o_objs_dict[propA][group][propB][subgroup].update({propC: type_formatter(old_dict[key], schema_properties, propA, propB, propC)})
									else:
										array_o_objs_dict[propA][group][propB][subgroup] = {propC: type_formatter(old_dict[key], schema_properties, propA, propB, propC)}
								else:
									array_o_objs_dict[propA][group][propB] = {}
									array_o_objs_dict[propA][group][propB][subgroup] = {propC: type_formatter(old_dict[key], schema_properties, propA, propB, propC)}
							else:
								array_o_objs_dict[propA][group] = {}
								array_o_objs_dict[propA][group][propB] = {}
								array_o_objs_dict[propA][group][propB][subgroup] = {propC: type_formatter(old_dict[key], schema_properties, propA, propB, propC)}
						else:
							array_o_objs_dict[propA] = {}
							array_o_objs_dict[propA][group] = {}
							array_o_objs_dict[propA][group][propB] = {}
							array_o_objs_dict[propA][group][propB][subgroup] = {propC: type_formatter(old_dict[key], schema_properties, propA, propB, propC)}
					else: # this is an object within an array of objects
						if schema_properties[propA]['items']['properties'][propB].get('linkTo') == 'OntologyTerm':
							if onts.get(propB):
								onts[propB][propC] = type_formatter(old_dict[key], ont_schema_properties, propC)
							else:
								onts[propB] = {}
								onts[propB][propC] = type_formatter(old_dict[key], ont_schema_properties, propC)
							if propC == 'term_id':
								if array_o_objs_dict.get(propA):
									if array_o_objs_dict[propA].get(group):
										array_o_objs_dict[propA][group][propB] = old_dict[key].replace(':','_') # replace the term_name/term_id with the linkTo
									else:
										array_o_objs_dict[propA][group] = {}
										array_o_objs_dict[propA][group][propB] = old_dict[key].replace(':','_') # replace the term_name/term_id with the linkTo
								else:
									array_o_objs_dict[propA] = {}
									array_o_objs_dict[propA][group] = {}
									array_o_objs_dict[propA][group][propB] = old_dict[key].replace(':','_') # replace the term_name/term_id with the linkTo
						else:
							print('ATTENTION IS NEEDED HERE her here her')
				else:
					if schema_properties[propA][propB]['type'] == 'array': # this is an array of objects within an object
						print('ATTENTION IS NEEDED HERE')
					else: # this is an object within an object
						if new_dict.get(propA): # I have already added the embedded object to the new dictionary
							if new_dict[propA].get(propB): # I have already added the double-embedded object to the new dictionary
								new_dict[propA][propB].update({propC: type_formatter(old_dict[key], schema_properties, propA, propB, propC)})
							else: # I need to add the double-embedded object to the new dictionary
								temp_dict = {propC: type_formatter(old_dict[key], schema_properties, propA, propB, propC)}
								new_dict[propA].update({propB: temp_dict})
						else: # I need to make new item in dictionary
							temp_dict = {propC: type_formatter(old_dict[key], schema_properties, propA, propB, propC)}
							new_dict[propA] = {propB: temp_dict}
			elif len(path) > 3:
				print('ERROR:' + key + ': Not prepared to do triplpe-embedded properties, check for errant period')

	for prop in array_o_objs_dict.keys():
		new_list = []
		for group_id in array_o_objs_dict[prop].keys():
			for i in array_o_objs_dict[prop][group_id].keys():
				if isinstance(array_o_objs_dict[prop][group_id][i], dict):
					nn_list = []
					for g_id in array_o_objs_dict[prop][group_id][i].keys():
						nn_list.append(array_o_objs_dict[prop][group_id][i][g_id])
					array_o_objs_dict[prop][group_id][i] = nn_list
			new_list.append(array_o_objs_dict[prop][group_id])
		new_dict[prop] = new_list
	return new_dict, onts


def attachment(path):
	''' Create an attachment upload object from a filename
	Embeds the attachment as a data url.
	'''
	if not os.path.isfile(path):
		r = requests.get(path)
		path = path.split('/')[-1]
		with open(path, 'wb') as outfile:
			outfile.write(r.content)
	filename = os.path.basename(path)
	mime_type, encoding = mimetypes.guess_type(path)
	major, minor = mime_type.split('/')
	detected_type = magic.from_file(path, mime=True)

	# XXX This validation logic should move server-side.
	if not (detected_type == mime_type or
			detected_type == 'text/plain' and major == 'text'):
		raise ValueError('Wrong extension for %s: %s' %
						 (detected_type, filename))

	with open(path, 'rb') as stream:
		attach = {
			'download': filename,
			'type': mime_type,
			'href': 'data:%s;base64,%s' % (mime_type, b64encode(stream.read()).decode('ascii'))
		}

		if mime_type in ('application/pdf', 'text/plain', 'text/tab-separated-values', 'text/html'):
			# XXX Should use chardet to detect charset for text files here.
			return attach

		if major == 'image' and minor in ('png', 'jpeg', 'gif', 'tiff'):
			# XXX we should just convert our tiffs to pngs
			stream.seek(0, 0)
			im = Image.open(stream)
			im.verify()
			if im.format != minor.upper():
				msg = 'Image file format %r does not match extension for %s'
				raise ValueError(msg % (im.format, filename))

			attach['width'], attach['height'] = im.size
			return attach

	raise ValueError('Unknown file type for %s' % filename)


def check_existing_obj(post_json, schema, connection):
	if post_json.get('uuid'):
		temp_identifier = post_json['uuid']
	elif post_json.get('accession'):
		temp_identifier = post_json['accession']
	elif post_json.get('external_accession'):
		temp_identifier = post_json['external_accession']
	elif post_json.get('name'):
		temp_identifier = schema + '/' + post_json['name']
	elif post_json.get('term_id'):
		temp_identifier = schema + '/' + post_json['term_id'].replace(':','_')
	elif post_json.get('gene_id'):
		temp_identifier = schema + '/' + post_json['gene_id']
	elif post_json.get('aliases'):
		temp_identifier = quote(post_json['aliases'][0])
	try:
		temp_identifier
	except NameError:
		temp = {}
		temp_identifier = ''
	else:
		url = urljoin(connection.server, temp_identifier + '/?format=json')
		temp = requests.get(url, auth=connection.auth).json()
	return temp_identifier, temp


def main():
	summary_report = []
	args = getArgs()
	if not args.mode:
		sys.exit('ERROR: --mode is required')
	connection = lattice.Connection(args.mode)
	server = connection.server
	print('Running on {server}'.format(server=server))
	if not os.path.isfile(args.infile):
		sys.exit('ERROR: file {filename} not found!'.format(filename=args.infile))
	book = load_workbook(args.infile)
	names = {}
	if args.justtype:
		if args.starttype:
			sys.exit('ERROR: cannot specify both --justtype and --starttype')
		else:
			for sheet in book.sheetnames:
				if sheet.lower().replace('_','') == args.justtype.lower().replace('_',''):
					names[sheet.lower().replace('_','')] = sheet
	else:
		for sheet in book.sheetnames:
			names[sheet.lower().replace('_','')] = sheet
	profiles = requests.get(server + 'profiles/?format=json').json()
	supported_collections = [s.lower() for s in list(profiles.keys())] # get accepted object types
	supported_collections.append('cover sheet')
	for n in names.keys():
		if n not in supported_collections: # check that each sheet name corresponds to an object type
			print('ERROR: Sheet name {name} not part of supported object types!'.format(
				name=n), file=sys.stderr)

	ont_schema_url = urljoin(server, 'profiles/ontology_term/?format=json')
	ont_term_schema = requests.get(ont_schema_url).json()['properties']
	ontology_props = []
	for p in ont_term_schema.keys():
		if not str(ont_term_schema[p].get('comment')).startswith('Do not submit') \
		and ont_term_schema[p].get('notSubmittable') != True:
			ontology_props.append(p)

	load_order = ORDER # pull in the order used to load test inserts on a local instance
	if args.starttype:
		st_index = load_order.index(args.starttype)
		load_order = load_order[st_index:]
	all_posts = {}
	for schema_to_load in load_order: # go in order to try and get objects posted before they are referenced by another object
		obj_type = schema_to_load.replace('_','')
		if obj_type in names.keys():
			obj_posts = []
			row_count, rows = reader(book, names[obj_type])

			# remove all columns that do not have any values submitted
			if not args.remove:
				index_to_remove = []
				for i in range(0,len(rows[0])):
					values = [row[i] for row in rows[1:]]
					if set(values) == {''}:
						index_to_remove.append(i)
				index_to_remove.reverse()
				for index in index_to_remove:
					for row in rows:
						del row[index]

			headers = rows.pop(0)
			schema_url = urljoin(server, 'profiles/' + schema_to_load + '/?format=json')
			schema_properties = requests.get(schema_url).json()['properties']
			invalid_flag = properties_validator(headers, schema_to_load, schema_properties, ontology_props)
			if invalid_flag == True:
				print('{}: invalid schema, check the headers'.format(obj_type))
				summary_report.append('{}: invalid schema, check the headers'.format(obj_type))
				continue
			for row in rows:
				row_count += 1
				post_json = dict(zip(headers, row))
				# convert values to the type specified in the schema, including embedded json objects
				if not args.remove:
					post_json, post_ont = dict_patcher(post_json, schema_properties, ont_term_schema)
					for k, v in post_ont.items():
						all_posts.setdefault('ontology_term', []).append((obj_type + '.' + k, v))
					# add attchments here
					if post_json.get('attachment'):
						attach = attachment(post_json['attachment'])
						post_json['attachment'] = attach
				obj_posts.append((row_count, post_json))
			all_posts[schema_to_load] = obj_posts

	if args.patchall:
		patch_req = True
	else:
		patch_req = False
	for schema in load_order: # go in order to try and get objects posted before they are referenced by another object
		if all_posts.get(schema):
			total = 0
			error = 0
			success = 0
			patch = 0
			new_accessions_aliases = []
			failed_postings = []
			for row_count, post_json in all_posts[schema]:
				total += 1
				#check for an existing object based on any possible identifier
				temp_id, temp = check_existing_obj(post_json, schema, connection)

				if temp.get('uuid'): # if there is an existing corresponding object
					if schema == 'ontology_term':
						ont_mismatch = False
						ont_patch = False
						for k in post_json.keys():
							if temp.get(k) and post_json[k] != temp.get(k):
								print('ERROR: {}:{} {} of {} does not match existing {}'.format(row_count, k, post_json[k], post_json['term_id'], temp.get(k)))
								ont_mismatch = True
							elif not temp.get(k):
								ont_patch = True
						if ont_mismatch == False and ont_patch == True:
							print(schema.upper() + ' ' + str(row_count) + ':Object {} already exists.  Would you like to patch it instead?'.format(post_json['term_id']))
							i = input('PATCH? y/n: ')
							if i.lower() == 'y':
						 		patch_req = True
						elif ont_mismatch == True:
							print('OntologyTerm {} will not be updated'.format(post_json['term_id']))
							i = input('EXIT SUBMISSION? y/n: ')
							if i.lower() == 'y':
								sys.exit('{sheet}: {success} posted, {patch} patched, {error} errors out of {total} total'.format(
									sheet=schema.upper(), success=success, total=total, error=error, patch=patch))
					elif patch_req == False: # patch wasn't specified, see if the user wants to patch
						print(schema.upper() + ' ROW ' + str(row_count) + ':Object {} already exists.  Would you like to patch it instead?'.format(temp_id))
						i = input('PATCH? y/n: ')
						if i.lower() == 'y':
							patch_req = True
					if patch_req == True and args.remove:
						existing_json = lattice.get_object(temp['uuid'], connection, frame="edit")
						for k in post_json.keys():
							if k not in ['uuid', 'accession', 'alias', '@id']:
								if k not in existing_json.keys():
									print('Cannot remove {}, may be calculated property, or is not submitted'.format(k))
								else:
									existing_json.pop(k)
									print('Removing value:', k)
						if args.update:
							e = lattice.replace_object(temp['uuid'], connection, existing_json)
							if e['status'] == 'error':
								error += 1
							elif e['status'] == 'success':
								new_patched_object = e['@graph'][0]
								# Print now and later
								print(schema.upper() + ' ROW ' + str(row_count) + ':identifier: {}'.format((new_patched_object.get(
									'accession', new_patched_object.get('uuid')))))
								patch += 1
					elif patch_req == True and args.update:
						e = lattice.patch_object(temp['uuid'], connection, post_json)
						if e['status'] == 'error':
							error += 1
						elif e['status'] == 'success':
							new_patched_object = e['@graph'][0]
							# Print now and later
							print(schema.upper() + ' ROW ' + str(row_count) + ':identifier: {}'.format((new_patched_object.get(
								'accession', new_patched_object.get('uuid')))))
							patch += 1
				else: # we have new object to post
					if args.patchall:
						print(schema.upper() + ' ROW ' + str(row_count) + ':Object not found. Check identifier or consider removing --patchall to post a new object')
						error += 1
					elif args.update:
						print(schema.upper() + ' ROW ' + str(row_count) + ':POSTing data!')
						e = lattice.post_object(schema, connection, post_json)
						if e['status'] == 'error':
							error += 1
							failed_postings.append(schema.upper() + ' ROW ' + str(row_count) + ':' + str(post_json.get(
								'aliases', 'alias not specified')))
						elif e['status'] == 'success':
							new_object = e['@graph'][0]
							# Print now and later
							print(schema.upper() + ' ROW ' + str(row_count) + ':New accession/UUID: {}'.format((new_object.get(
								'accession', new_object.get('uuid')))))
							new_accessions_aliases.append(('ROW ' + str(row_count), new_object.get(
								'accession', new_object.get('uuid')), new_object.get('aliases', new_object.get('name'))))
							success += 1

			# Print now and later
			print('{sheet}: {success} posted, {patch} patched, {error} errors out of {total} total'.format(
				sheet=schema.upper(), success=success, total=total, error=error, patch=patch))
			summary_report.append('{sheet}: {success} posted, {patch} patched, {error} errors out of {total} total'.format(
				sheet=schema.upper(), success=success, total=total, error=error, patch=patch))
			if new_accessions_aliases:
				print('New accessions/UUIDs and aliases:')
				for (row, accession, alias) in new_accessions_aliases:
					if alias == None:
						alias = 'alias not specified'
					else:
						alias = ', '.join(alias) if isinstance(alias, list) else alias
					print(row, accession, alias)
			if failed_postings:
				print('Posting failed for {} object(s):'.format(len(failed_postings)))
				for alias in failed_postings:
					print(', '.join(alias) if isinstance(alias, list) else alias)
	print('-------Summary of all objects-------')
	print('\n'.join(summary_report))


if __name__ == '__main__':
	main()
